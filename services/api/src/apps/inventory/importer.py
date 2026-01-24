from __future__ import annotations

from decimal import Decimal, InvalidOperation
from typing import Any, Iterable, List

from django.core.exceptions import ValidationError
from django.db import transaction
from openpyxl import load_workbook

from apps.business.models import Business
from apps.catalog.models import Product
from .models import InventoryImportJob
from .services import ensure_stock_record, register_stock_movement

MAX_ROWS = 2000
PREVIEW_LIMIT = 200
HEADER_ALIASES = {
		'nombre': 'name',
		'producto': 'name',
		'sku': 'sku',
		'codigo': 'sku',
		'código': 'sku',
		'barcode': 'barcode',
		'codigo de barras': 'barcode',
		'código de barras': 'barcode',
		'precio': 'price',
		'precio venta': 'price',
		'costo': 'cost',
		'stock minimo': 'stock_min',
		'stock mínimo': 'stock_min',
		'stock_min': 'stock_min',
		'stock': 'stock',
		'nota': 'note',
		'notas': 'note',
}
REQUIRED_FIELDS = {'name'}
SUPPORTED_FIELDS = {'name', 'sku', 'barcode', 'price', 'cost', 'stock', 'stock_min', 'note'}
DEFAULT_NOTE = 'Importación de stock (.xlsx)'


class InventoryImportError(Exception):
	"""Error amigable para el importador de inventario."""


def parse_inventory_import(file_obj, *, business: Business) -> tuple[list[dict[str, Any]], dict[str, int]]:
	file_obj.seek(0)
	try:
		workbook = load_workbook(filename=file_obj, data_only=True)
	except Exception as exc:  # pragma: no cover - openpyxl lanza varias subclases
		raise InventoryImportError('No pudimos leer el archivo. Asegurate de que sea un .xlsx válido.') from exc

	sheet = workbook.active
	header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
	if not header_row:
		raise InventoryImportError('El archivo está vacío. Descargá la plantilla más reciente e intentá de nuevo.')

	header_map = _build_header_map(header_row)
	if not all(field in header_map.values() for field in REQUIRED_FIELDS):
		raise InventoryImportError('La plantilla debe incluir la columna "Nombre".')

	products = Product.objects.filter(business=business)
	sku_lookup = {p.sku.lower(): p for p in products if p.sku}
	name_lookup = {p.name.strip().lower(): p for p in products}
	seen_new_skus: set[str] = set()

	rows: list[dict[str, Any]] = []
	summary = {
		'total_rows': 0,
		'create_count': 0,
		'update_count': 0,
		'adjust_count': 0,
		'skip_count': 0,
		'warning_count': 0,
		'error_count': 0,
	}

	processed = 0
	for line_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
		if processed >= MAX_ROWS:
			break
		normalized_row = _normalize_row(values, header_map)
		if not any(_has_value(value) for value in normalized_row.values()):
			continue
		processed += 1
		summary['total_rows'] += 1

		result_row, status = _build_row(
			normalized_row,
			line_number=line_number,
			sku_lookup=sku_lookup,
			name_lookup=name_lookup,
			seen_new_skus=seen_new_skus,
		)
		rows.append(result_row)

		if status == 'error':
			summary['error_count'] += 1
		elif status == 'warning':
			summary['warning_count'] += 1

		if status != 'error':
			if result_row['action'] == 'create':
				summary['create_count'] += 1
			else:
				summary['update_count'] += 1
			if result_row['stock_action'] == 'adjust':
				summary['adjust_count'] += 1
			else:
				summary['skip_count'] += 1
		else:
			if result_row['stock_action'] == 'skip':
				summary['skip_count'] += 1

	if summary['total_rows'] == 0:
		raise InventoryImportError('No encontramos filas con datos. Revisá que la hoja tenga información desde la fila 2 en adelante.')

	return rows, summary


def apply_inventory_import(job: InventoryImportJob, *, business: Business, user) -> InventoryImportJob:
	if job.business_id != business.id:
		raise InventoryImportError('No encontramos esta importación en tu negocio actual.')
	if job.summary and job.summary.get('error_count'):
		raise InventoryImportError('Corregí los errores del archivo antes de aplicar la importación.')

	created = updated = adjusted = skipped = 0
	invalid_rows = 0
	cache_by_id: dict[str, Product] = {}

	try:
		with transaction.atomic():
			job.status = InventoryImportJob.Status.PROCESSING
			job.save(update_fields=['status', 'updated_at'])

			for row in job.rows:
				if row.get('status') == 'error':
					invalid_rows += 1
					continue

				payload = row.get('payload') or {}
				product = _resolve_product(row, business, cache_by_id)
				if product is None:
					product = _create_product(payload, business)
					cache_by_id[str(product.id)] = product
					created += 1
				else:
					_update_product(product, payload)
					updated += 1

				stock_value = _to_decimal(payload.get('stock'))
				if stock_value is not None:
					register_stock_movement(
						business=business,
						product=product,
						movement_type='ADJUST',
						quantity=stock_value,
						note=payload.get('note') or DEFAULT_NOTE,
						created_by=user if getattr(user, 'is_authenticated', False) else None,
					)
					adjusted += 1
				else:
					skipped += 1

			job.status = InventoryImportJob.Status.DONE
			job.created_count = created
			job.updated_count = updated
			job.adjusted_count = adjusted
			job.skipped_count = skipped
			job.error_count = job.summary.get('error_count', invalid_rows) if job.summary else invalid_rows
			job.warning_count = job.summary.get('warning_count', 0) if job.summary else 0
			job.save(update_fields=[
				'status',
				'created_count',
				'updated_count',
				'adjusted_count',
				'skipped_count',
				'error_count',
				'warning_count',
				'updated_at',
			])
	except (InventoryImportError, ValidationError) as exc:
		_mark_job_as_failed(job, str(exc))
		if isinstance(exc, InventoryImportError):
			raise
		raise InventoryImportError(str(exc)) from exc
	except Exception as exc:  # pragma: no cover - fallback defensivo
		_mark_job_as_failed(job, 'Error inesperado al aplicar la importación.')
		raise InventoryImportError('Ocurrió un error inesperado al aplicar la importación.') from exc

	return job


def serialize_preview_rows(rows: list[dict[str, Any]], *, limit: int = PREVIEW_LIMIT) -> list[dict[str, Any]]:
	sanitized: list[dict[str, Any]] = []
	for row in rows[:limit]:
		public_row = {k: v for k, v in row.items() if k not in {'payload', 'product_id'}}
		sanitized.append(public_row)
	return sanitized


def _build_header_map(header_row: Iterable[Any]) -> dict[int, str]:
	mapping: dict[int, str] = {}
	for index, raw_value in enumerate(header_row, start=1):
		normalized = _normalize_header(raw_value)
		if not normalized:
			continue
		field = HEADER_ALIASES.get(normalized, normalized)
		if field in SUPPORTED_FIELDS:
			mapping[index] = field
	return mapping


def _normalize_row(values: Iterable[Any], header_map: dict[int, str]) -> dict[str, Any]:
	normalized: dict[str, Any] = {}
	for index, value in enumerate(values, start=1):
		field = header_map.get(index)
		if field:
			normalized[field] = value
	return normalized


def _build_row(
	normalized_row: dict[str, Any],
	*,
	line_number: int,
	sku_lookup: dict[str, Product],
	name_lookup: dict[str, Product],
	seen_new_skus: set[str],
) -> tuple[dict[str, Any], str]:
	messages: list[str] = []
	status = 'ok'

	name = _clean_string(normalized_row.get('name'))
	if not name:
		messages.append('El campo "Nombre" es obligatorio.')
	elif len(name) > 255:
		messages.append('El nombre supera los 255 caracteres permitidos.')

	raw_sku = _clean_string(normalized_row.get('sku')) or None
	if raw_sku and len(raw_sku) > 64:
		messages.append('El SKU supera los 64 caracteres permitidos.')
	sku_key = raw_sku.lower() if raw_sku else None

	barcode = _clean_string(normalized_row.get('barcode')) or None
	if barcode and len(barcode) > 128:
		messages.append('El código de barras supera los 128 caracteres permitidos.')

	price, price_error = _parse_decimal(normalized_row.get('price'), field_label='Precio')
	cost, cost_error = _parse_decimal(normalized_row.get('cost'), field_label='Costo')
	stock_min, stock_min_error = _parse_decimal(normalized_row.get('stock_min'), field_label='Stock mínimo')
	stock, stock_error = _parse_decimal(normalized_row.get('stock'), field_label='Stock')

	for error in [price_error, cost_error, stock_min_error, stock_error]:
		if error:
			messages.append(error)

	note = _clean_string(normalized_row.get('note')) or DEFAULT_NOTE

	product = None
	if sku_key and sku_key in sku_lookup:
		product = sku_lookup[sku_key]
	elif name.strip().lower() in name_lookup:
		product = name_lookup[name.strip().lower()]
	elif sku_key:
		if sku_key in seen_new_skus:
			messages.append('Este SKU está repetido en el archivo.')
		else:
			seen_new_skus.add(sku_key)

	if messages:
		status = 'error'

	action = 'update' if product else 'create'
	stock_action = 'adjust' if stock is not None else 'skip'

	payload = {
		'name': name,
		'sku': raw_sku,
		'barcode': barcode,
		'price': _serialize_decimal(price),
		'cost': _serialize_decimal(cost),
		'stock_min': _serialize_decimal(stock_min),
		'stock': _serialize_decimal(stock),
		'note': note,
	}

	preview_values = {
		'name': name,
		'sku': raw_sku,
		'barcode': barcode,
		'price': _display_decimal(price),
		'cost': _display_decimal(cost),
		'stock_min': _display_decimal(stock_min),
		'stock': _display_decimal(stock),
	}

	row = {
		'line_number': line_number,
		'name': name,
		'sku': raw_sku,
		'action': action,
		'stock_action': stock_action,
		'status': status,
		'messages': messages,
		'values': preview_values,
		'payload': payload,
		'product_id': str(product.id) if product else None,
	}
	return row, status


def _resolve_product(row: dict[str, Any], business: Business, cache_by_id: dict[str, Product]):
	product_id = row.get('product_id')
	if not product_id:
		return None
	if product_id in cache_by_id:
		return cache_by_id[product_id]
	try:
		product = Product.objects.get(pk=product_id, business=business)
	except Product.DoesNotExist:
		return None
	cache_by_id[product_id] = product
	return product


def _create_product(payload: dict[str, Any], business: Business) -> Product:
	if not payload.get('name'):
		raise InventoryImportError('No podemos crear un producto sin nombre.')
	product = Product.objects.create(
		business=business,
		name=payload.get('name'),
		sku=payload.get('sku') or '',
		barcode=payload.get('barcode') or '',
		price=_to_decimal(payload.get('price')) or Decimal('0'),
		cost=_to_decimal(payload.get('cost')) or Decimal('0'),
		stock_min=_to_decimal(payload.get('stock_min')) or Decimal('0'),
	)
	ensure_stock_record(business, product)
	return product


def _update_product(product: Product, payload: dict[str, Any]) -> None:
	updates: List[str] = []
	for field in ['name', 'sku', 'barcode']:
		value = payload.get(field)
		if value is not None and value != getattr(product, field):
			setattr(product, field, value)
			updates.append(field)

	for field in ['price', 'cost', 'stock_min']:
		raw_value = payload.get(field)
		decimal_value = _to_decimal(raw_value)
		if decimal_value is not None and decimal_value != getattr(product, field):
			setattr(product, field, decimal_value)
			updates.append(field)

	if updates:
		product.save(update_fields=updates + ['updated_at'])


def _parse_decimal(value: Any, *, field_label: str) -> tuple[Decimal | None, str | None]:
	if value in (None, ''):
		return None, None
	try:
		decimal_value = Decimal(str(value))
	except (InvalidOperation, TypeError, ValueError):
		return None, f'{field_label} debe ser un número válido.'
	if decimal_value < 0:
		return None, f'{field_label} debe ser mayor o igual a cero.'
	return decimal_value, None


def _to_decimal(value: Any) -> Decimal | None:
	if value in (None, '', 'null'):
		return None
	try:
		return Decimal(str(value))
	except (InvalidOperation, TypeError, ValueError):
		return None


def _serialize_decimal(value: Decimal | None) -> str | None:
	if value is None:
		return None
	return format(value, 'f')


def _display_decimal(value: Decimal | None) -> float | None:
	if value is None:
		return None
	return float(value)


def _normalize_header(raw_value: Any) -> str:
	if raw_value is None:
		return ''
	return str(raw_value).strip().lower()


def _clean_string(value: Any) -> str:
	if value is None:
		return ''
	return str(value).strip()


def _has_value(value: Any) -> bool:
	if value is None:
		return False
	if isinstance(value, str) and not value.strip():
		return False
	return True


def _mark_job_as_failed(job: InventoryImportJob, message: str) -> None:
	errors = list(job.errors or [])
	errors.append(message)
	job.status = InventoryImportJob.Status.FAILED
	job.errors = errors
	job.save(update_fields=['status', 'errors', 'updated_at'])
