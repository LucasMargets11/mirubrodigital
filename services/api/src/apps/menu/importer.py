from __future__ import annotations

from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any, Iterable

from django.db import transaction
from openpyxl import Workbook, load_workbook

from apps.business.models import Business
from .models import MenuCategory, MenuItem

MAX_ROWS = 1200
PREVIEW_LIMIT = 200
HEADER_ALIASES = {
    'categoria': 'category',
    'categoría': 'category',
    'category': 'category',
    'nombre': 'name',
    'name': 'name',
    'descripcion': 'description',
    'descripción': 'description',
    'description': 'description',
    'precio': 'price',
    'price': 'price',
    'sku': 'sku',
    'codigo': 'sku',
    'código': 'sku',
    'tags': 'tags',
    'etiquetas': 'tags',
    'disponible': 'is_available',
    'available': 'is_available',
    'activo': 'is_available',
    'destacado': 'is_featured',
    'featured': 'is_featured',
    'orden': 'position',
    'position': 'position',
    'posicion': 'position',
    'posición': 'position',
    'tiempo': 'estimated_time',
    'tiempo prep': 'estimated_time',
}
TRUE_VALUES = {'si', 'sí', 'true', '1', 'x', 'yes'}
FALSE_VALUES = {'no', 'false', '0'}


class MenuImportError(Exception):
    """Error amigable para el importador de carta."""


def apply_menu_import(file_obj, *, business: Business) -> dict[str, Any]:
    file_obj.seek(0)
    try:
        workbook = load_workbook(filename=file_obj, data_only=True)
    except Exception as exc:  # pragma: no cover - openpyxl lanza varias subclases
        raise MenuImportError('No pudimos leer el archivo. Confirmá que sea un .xlsx válido.') from exc

    sheet = workbook.active
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise MenuImportError('El archivo está vacío. Descargá la última plantilla e intentá nuevamente.')

    header_map = _build_header_map(header_row)
    if 'name' not in header_map.values():
        raise MenuImportError('La plantilla debe incluir una columna "Nombre".')

    summary = {
        'total_rows': 0,
        'created_categories': 0,
        'updated_categories': 0,
        'created_items': 0,
        'updated_items': 0,
        'skipped_rows': 0,
    }
    preview: list[dict[str, Any]] = []

    categories_by_name = {
        cat.name.strip().lower(): cat
        for cat in MenuCategory.objects.filter(business=business)
    }
    items_by_sku = {
        item.sku.strip().lower(): item
        for item in MenuItem.objects.filter(business=business).exclude(sku='')
    }
    items_by_name = {
        (item.category_id, item.name.strip().lower()): item
        for item in MenuItem.objects.filter(business=business)
    }

    processed = 0

    with transaction.atomic():
        for line_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if processed >= MAX_ROWS:
                break
            normalized = _normalize_row(values, header_map)
            if not _has_values(normalized):
                continue
            processed += 1
            summary['total_rows'] += 1

            try:
                result = _process_row(
                    normalized,
                    line_number=line_number,
                    business=business,
                    categories_by_name=categories_by_name,
                    items_by_sku=items_by_sku,
                    items_by_name=items_by_name,
                )
            except MenuImportError as exc:
                raise exc

            if result is None:
                summary['skipped_rows'] += 1
                continue

            if result['category_action'] == 'create':
                summary['created_categories'] += 1
            elif result['category_action'] == 'update':
                summary['updated_categories'] += 1

            if result['item_action'] == 'create':
                summary['created_items'] += 1
            elif result['item_action'] == 'update':
                summary['updated_items'] += 1

            if len(preview) < PREVIEW_LIMIT:
                preview.append(
                    {
                        'line_number': line_number,
                        'category': result['category_name'],
                        'name': result['item_name'],
                        'price': result['price_display'],
                        'action': result['item_action'],
                        'available': result['is_available'],
                    }
                )

    return {'summary': summary, 'preview': preview}


def export_menu_to_workbook(*, business: Business) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Carta'
    sheet.append(['Categoría', 'Nombre', 'Descripción', 'Precio', 'Disponible', 'Destacado', 'Tags', 'Orden', 'Tiempo'])

    items = (
        MenuItem.objects.filter(business=business)
        .select_related('category')
        .order_by('category__position', 'category__name', 'position', 'name')
    )
    for item in items:
        sheet.append(
            [
                item.category.name if item.category else '',
                item.name,
                item.description,
                float(item.price),
                'Sí' if item.is_available else 'No',
                'Sí' if item.is_featured else 'No',
                ', '.join(item.tag_list),
                item.position,
                item.estimated_time_minutes,
            ]
        )

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.read()


def _process_row(
    row: dict[str, Any],
    *,
    line_number: int,
    business: Business,
    categories_by_name: dict[str, MenuCategory],
    items_by_sku: dict[str, MenuItem],
    items_by_name: dict[tuple[str | None, str], MenuItem],
):
    name = _clean_string(row.get('name'))
    if not name:
        return None
    if len(name) > 255:
        raise MenuImportError(f'Fila {line_number}: el nombre supera los 255 caracteres permitidos.')

    category_name = _clean_string(row.get('category'))
    category_key = category_name.lower() if category_name else ''
    category_action = 'skip'
    if category_key:
        category = categories_by_name.get(category_key)
        if category is None:
            category = MenuCategory.objects.create(
                business=business,
                name=category_name,
                description='',
                position=len(categories_by_name) + 1,
            )
            categories_by_name[category_key] = category
            category_action = 'create'
        else:
            category_action = 'skip'
    else:
        category = None

    description = _clean_string(row.get('description'))
    price = _parse_decimal(row.get('price'), field_label='Precio', line_number=line_number)
    sku = _clean_string(row.get('sku'))
    tags = _normalize_tags(row.get('tags'))
    is_available = _parse_bool(row.get('is_available'), default=True, field_label='Disponible', line_number=line_number)
    is_featured = _parse_bool(row.get('is_featured'), default=False, field_label='Destacado', line_number=line_number)
    position = _parse_int(row.get('position'), field_label='Orden', line_number=line_number)
    estimated_time = _parse_int(row.get('estimated_time'), field_label='Tiempo', line_number=line_number)

    item = None
    item_action = 'create'
    lookup_key = None
    if sku:
        lookup_key = sku.lower()
        item = items_by_sku.get(lookup_key)
    if item is None:
        name_key = name.strip().lower()
        lookup_key = (category.id if category else None, name_key)
        item = items_by_name.get(lookup_key)

    payload = {
        'business': business,
        'category': category,
        'name': name,
        'description': description,
        'price': price,
        'sku': sku,
        'tags': tags,
        'is_available': is_available,
        'is_featured': is_featured,
        'position': position,
        'estimated_time_minutes': estimated_time,
    }

    if item is None:
        item = MenuItem.objects.create(**payload)
        if sku:
            items_by_sku[sku.lower()] = item
        items_by_name[(item.category_id, item.name.strip().lower())] = item
        item_action = 'create'
    else:
        updates = []
        for field, value in payload.items():
            if field == 'business':
                continue
            if getattr(item, field) != value:
                setattr(item, field, value)
                updates.append(field)
        if updates:
            item.save(update_fields=updates + ['updated_at'])
            item_action = 'update'
        else:
            item_action = 'skip'

    if lookup_key and isinstance(lookup_key, str):
        items_by_sku[lookup_key] = item
    else:
        items_by_name[(item.category_id, item.name.strip().lower())] = item

    return {
        'category_action': category_action,
        'item_action': item_action,
        'category_name': category.name if category else '',
        'item_name': item.name,
        'price_display': f"{item.price:.2f}",
        'is_available': item.is_available,
    }


def _build_header_map(header_row: Iterable[Any]) -> dict[int, str]:
    mapping: dict[int, str] = {}
    for index, raw_value in enumerate(header_row, start=1):
        normalized = _clean_string(raw_value).lower()
        if not normalized:
            continue
        field = HEADER_ALIASES.get(normalized, normalized)
        mapping[index] = field
    return mapping


def _normalize_row(values: Iterable[Any], header_map: dict[int, str]) -> dict[str, Any]:
    normalized: dict[str, Any] = {}
    for index, value in enumerate(values, start=1):
        field = header_map.get(index)
        if field:
            normalized[field] = value
    return normalized


def _has_values(row: dict[str, Any]) -> bool:
    return any(value not in (None, '', []) for value in row.values())


def _clean_string(value: Any) -> str:
    if value in (None, ''):
        return ''
    return str(value).strip()


def _normalize_tags(value: Any) -> str:
    if value in (None, ''):
        return ''
    if isinstance(value, str):
        raw_tags = value.split(',')
    elif isinstance(value, Iterable):
        raw_tags = value
    else:
        raw_tags = []
    normalized = []
    seen = set()
    for tag in raw_tags:
        clean = str(tag).strip()
        if not clean:
            continue
        lowered = clean.lower()
        if lowered in seen:
            continue
        seen.add(lowered)
        normalized.append(clean)
    return ','.join(normalized)


def _parse_decimal(value: Any, *, field_label: str, line_number: int) -> Decimal:
    if value in (None, ''):
        return Decimal('0')
    try:
        decimal_value = Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError) as exc:
        raise MenuImportError(f'Fila {line_number}: {field_label} debe ser un número válido.') from exc
    if decimal_value < 0:
        raise MenuImportError(f'Fila {line_number}: {field_label} debe ser mayor o igual a 0.')
    return decimal_value


def _parse_bool(value: Any, *, default: bool, field_label: str, line_number: int) -> bool:
    if value in (None, ''):
        return default
    normalized = str(value).strip().lower()
    if normalized in TRUE_VALUES:
        return True
    if normalized in FALSE_VALUES:
        return False
    raise MenuImportError(f'Fila {line_number}: {field_label} debe ser Sí/No.')


def _parse_int(value: Any, *, field_label: str, line_number: int) -> int:
    if value in (None, ''):
        return 0
    try:
        integer = int(str(value).strip())
    except (TypeError, ValueError) as exc:
        raise MenuImportError(f'Fila {line_number}: {field_label} debe ser un número entero.') from exc
    return max(0, integer)
